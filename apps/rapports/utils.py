"""
Utilitaires pour la génération de rapports
"""

from django.conf import settings
from django.db.models import Sum, Count, Q
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
import openpyxl
import csv
import os
from datetime import datetime
from dateutil.relativedelta import relativedelta

from apps.cotisations.models import Cotisation
from apps.paiements.models import Paiement


def generer_rapport_pdf(rapport, options=None):
    """Générer un rapport PDF"""
    if options is None:
        options = {}

    # Créer le répertoire s'il n'existe pas
    rapport_dir = os.path.join(settings.MEDIA_ROOT, 'rapports', str(rapport.periode.year),
                               f"{rapport.periode.month:02d}")
    os.makedirs(rapport_dir, exist_ok=True)

    # Nom du fichier
    nom_fichier = f"rapport_{rapport.type_rapport}_{rapport.association.id}_{rapport.periode.strftime('%Y-%m')}.pdf"
    fichier_path = os.path.join(rapport_dir, nom_fichier)

    # Collecter les données
    donnees = _collecter_donnees_rapport(rapport)
    rapport.set_donnees(donnees)

    # Créer le PDF
    c = canvas.Canvas(fichier_path, pagesize=A4)
    width, height = A4

    # En-tête
    c.setFont("Helvetica-Bold", 20)
    c.drawString(50, height - 80, f"Rapport {rapport.get_type_rapport_display()}")

    c.setFont("Helvetica", 14)
    c.drawString(50, height - 110, f"Association: {rapport.association.nom}")
    c.drawString(50, height - 130, f"Période: {rapport.periode.strftime('%B %Y')}")
    c.drawString(50, height - 150, f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}")

    # Données principales
    y = height - 200
    c.setFont("Helvetica-Bold", 16)
    c.drawString(50, y, "Résumé Financier")

    y -= 30
    c.setFont("Helvetica", 12)

    resume_items = [
        f"Total collecté: {donnees['total_collecte']} DA",
        f"Total attendu: {donnees['total_attendu']} DA",
        f"Taux de recouvrement: {donnees['taux_recouvrement']:.1f}%",
        f"Nombre de paiements: {donnees['nombre_paiements']}",
        f"Cotisations en retard: {donnees['cotisations_retard']}",
    ]

    for item in resume_items:
        c.drawString(50, y, item)
        y -= 20

    # Détails par logement si demandé
    if options.get('inclure_details', False):
        y -= 30
        c.setFont("Helvetica-Bold", 14)
        c.drawString(50, y, "Détails par Logement")
        y -= 20

        c.setFont("Helvetica", 10)
        for detail in donnees['details_logements'][:20]:  # Limiter à 20 pour ne pas déborder
            c.drawString(50, y,
                         f"{detail['logement']} - {detail['resident']} - {detail['statut']} - {detail['montant']} DA")
            y -= 15

            if y < 50:  # Nouvelle page si nécessaire
                c.showPage()
                y = height - 50

    c.save()

    return fichier_path


def generer_rapport_excel(rapport, options=None):
    """Générer un rapport Excel"""
    if options is None:
        options = {}

    # Créer le répertoire s'il n'existe pas
    rapport_dir = os.path.join(settings.MEDIA_ROOT, 'rapports', str(rapport.periode.year),
                               f"{rapport.periode.month:02d}")
    os.makedirs(rapport_dir, exist_ok=True)

    # Nom du fichier
    nom_fichier = f"rapport_{rapport.type_rapport}_{rapport.association.id}_{rapport.periode.strftime('%Y-%m')}.xlsx"
    fichier_path = os.path.join(rapport_dir, nom_fichier)

    # Collecter les données
    donnees = _collecter_donnees_rapport(rapport)
    rapport.set_donnees(donnees)

    # Créer le classeur Excel
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Résumé"

    # En-tête
    worksheet['A1'] = f"Rapport {rapport.get_type_rapport_display()}"
    worksheet['A2'] = f"Association: {rapport.association.nom}"
    worksheet['A3'] = f"Période: {rapport.periode.strftime('%B %Y')}"
    worksheet['A4'] = f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"

    # Données principales
    row = 6
    worksheet[f'A{row}'] = "Résumé Financier"
    row += 2

    resume_data = [
        ['Total collecté', f"{donnees['total_collecte']} DA"],
        ['Total attendu', f"{donnees['total_attendu']} DA"],
        ['Taux de recouvrement', f"{donnees['taux_recouvrement']:.1f}%"],
        ['Nombre de paiements', donnees['nombre_paiements']],
        ['Cotisations en retard', donnees['cotisations_retard']],
    ]

    for item in resume_data:
        worksheet[f'A{row}'] = item[0]
        worksheet[f'B{row}'] = item[1]
        row += 1

    # Détails par logement si demandé
    if options.get('inclure_details', False):
        # Créer une nouvelle feuille
        details_sheet = workbook.create_sheet(title="Détails Logements")

        # En-têtes
        headers = ['Logement', 'Résident', 'Statut', 'Montant', 'Date Échéance']
        for col, header in enumerate(headers, 1):
            details_sheet.cell(row=1, column=col, value=header)

        # Données
        for row, detail in enumerate(donnees['details_logements'], 2):
            details_sheet.cell(row=row, column=1, value=detail['logement'])
            details_sheet.cell(row=row, column=2, value=detail['resident'])
            details_sheet.cell(row=row, column=3, value=detail['statut'])
            details_sheet.cell(row=row, column=4, value=detail['montant'])
            details_sheet.cell(row=row, column=5, value=detail['echeance'])

    workbook.save(fichier_path)

    return fichier_path


def generer_rapport_csv(rapport, options=None):
    """Générer un rapport CSV"""
    if options is None:
        options = {}

    # Créer le répertoire s'il n'existe pas
    rapport_dir = os.path.join(settings.MEDIA_ROOT, 'rapports', str(rapport.periode.year),
                               f"{rapport.periode.month:02d}")
    os.makedirs(rapport_dir, exist_ok=True)

    # Nom du fichier
    nom_fichier = f"rapport_{rapport.type_rapport}_{rapport.association.id}_{rapport.periode.strftime('%Y-%m')}.csv"
    fichier_path = os.path.join(rapport_dir, nom_fichier)

    # Collecter les données
    donnees = _collecter_donnees_rapport(rapport)
    rapport.set_donnees(donnees)

    # Créer le CSV
    with open(fichier_path, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.writer(csvfile)

        # En-tête du rapport
        writer.writerow([f"Rapport {rapport.get_type_rapport_display()}"])
        writer.writerow([f"Association: {rapport.association.nom}"])
        writer.writerow([f"Période: {rapport.periode.strftime('%B %Y')}"])
        writer.writerow([f"Généré le: {datetime.now().strftime('%d/%m/%Y à %H:%M')}"])
        writer.writerow([])  # Ligne vide

        # Résumé financier
        writer.writerow(['Résumé Financier'])
        writer.writerow(['Total collecté', f"{donnees['total_collecte']} DA"])
        writer.writerow(['Total attendu', f"{donnees['total_attendu']} DA"])
        writer.writerow(['Taux de recouvrement', f"{donnees['taux_recouvrement']:.1f}%"])
        writer.writerow(['Nombre de paiements', donnees['nombre_paiements']])
        writer.writerow(['Cotisations en retard', donnees['cotisations_retard']])
        writer.writerow([])  # Ligne vide

        # Détails par logement
        if options.get('inclure_details', False):
            writer.writerow(['Détails par Logement'])
            writer.writerow(['Logement', 'Résident', 'Statut', 'Montant', 'Date Échéance'])

            for detail in donnees['details_logements']:
                writer.writerow([
                    detail['logement'],
                    detail['resident'],
                    detail['statut'],
                    detail['montant'],
                    detail['echeance']
                ])

    return fichier_path


def _collecter_donnees_rapport(rapport):
    """Collecter les données pour le rapport"""
    association = rapport.association
    periode = rapport.periode

    # Déterminer la période selon le type de rapport
    if rapport.type_rapport == 'mensuel':
        debut_periode = periode.replace(day=1)
        fin_periode = (debut_periode + relativedelta(months=1)) - relativedelta(days=1)
    elif rapport.type_rapport == 'trimestriel':
        # Premier jour du trimestre
        debut_periode = periode.replace(day=1, month=((periode.month - 1) // 3) * 3 + 1)
